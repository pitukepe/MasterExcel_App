import pandas as pd
import numpy as np
import os
import sqlite3
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt

#### Main Class ####
class App:
    def __init__(self, master):
        self.master = master
        self.main_page()

    ### Functions for creating the GUI and its pages ###

    ## Main page
    def main_page(self):
        # Destroying all widgets from the previous page (when returned into from another page)
        for i in self.master.winfo_children():
            i.destroy()

        # title and geometry
        self.master.title("Excel Master")
        self.master.geometry("280x300")
        self.master.config(bg="#0b2838")
        
        # greeting message and menu buttons
        tk.Label(self.master, text="Welcome to Excel Master!", font=("Times","24"), fg="#5ea832", bg="#0b2838").grid(row=0, column=0, padx=10, pady=10)
        tk.Label(self.master, text="Choose what you want to do:", font=("Times","15"), fg="#5ea832" ,bg="#0b2838").grid(row=1, column=0, padx=10)

        # Excel cleaner site button
        self.clean_button = tk.Button(self.master, text="Excel Cleaner", command=self.clean_page, activeforeground="blue", background="#0b2838")
        self.clean_button.grid(row=2, column=0, padx=10, pady=10)

        # Excel pivot creator site button
        self.pivot_button = tk.Button(self.master, text="Excel Pivot Creator", command = self.pivot_page, activeforeground="blue", background="#0b2838")
        self.pivot_button.grid(row=3, column=0, padx=10, pady=10)

        # Excel plotter site button
        self.plotter_button = tk.Button(self.master, text="Excel Graph Plotter", command = self.plotter_page, activeforeground="blue", background="#0b2838")
        self.plotter_button.grid(row=4, column=0, padx=10, pady=10)

        # Excel formula applier site button
        self.formula_button = tk.Button(self.master, text="Excel Formula Applier", command = self.formula_page, activeforeground="blue", background="#0b2838")
        self.formula_button.grid(row=5, column=0, padx=10, pady=10)
        
        # Trademark
        tk.Label(self.master, text="© Made by Peter Peško, 2024", font=("Times","12"), fg="#d4d4d4", bg="#0b2838").grid(row=6, column=0, padx=10, sticky="w")

    ## Cleaner page
    def clean_page(self):
        # Destroying all widgets from the main page
        for i in self.master.winfo_children():
            i.destroy()
        
        # Cleaner page geometry
        self.master.geometry("520x240")
        
        # Back button
        back_button = tk.Button(self.master, text="<<<", command=self.main_page, cursor="hand2", activeforeground="blue")
        back_button.grid(row=0, column=0, padx=10, sticky="w")

        # Cleaner page title
        tk.Label(self.master, text = "Excel Cleaner", font=("Times","20"), fg="#5ea832", bg="#0b2838").grid(row=1, column=1, padx=10, sticky="w")

        # Choosing an Excel file with button
        self.file_label = tk.Label(self.master, text="Choose Excel file:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.file_label.grid(row=2, column=0, padx=5, sticky="w")
        self.file = tk.Entry(self.master, state="disabled")
        self.file.grid(row=2, column=1, sticky="w")
        file_button = tk.Button(self.master, text="Choose file", command=self.choose_file, activeforeground="blue", bg="#0b2838")
        file_button.grid(row=2, column=2, sticky="w")

        # Specifying if there is an index in the excell file
        tk.Label(self.master, text="Does your file have index?", font=("Times","15"), fg="#5ea832", bg="#0b2838").grid(row=3, column=0, padx=5, sticky="w")
        self.index_check = tk.IntVar()
        self.index = tk.Checkbutton(self.master, variable=self.index_check, text="(no)", onvalue=1, offvalue=0, activeforeground="blue", command=self.toggle_index, bg="#0b2838")
        self.index.grid(row=3, column=1, sticky="w")
        self.index_col_label = tk.Label(self.master, text="Index col. number:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.index_col = tk.Entry(self.master, width=5)
    
        # Choosing preferred output file in a dropdown menu
        tk.Label(self.master, text="Preferred output:", font=("Times","15"), fg="#5ea832", bg="#0b2838").grid(row=4, column=0, padx=5,sticky="w")
        self.n = tk.StringVar()
        self.output_choice = ttk.Combobox(self.master, state="readonly", values=[".xlsx", ".csv", ".sqlite"], textvariable=self.n, width=10)
        self.output_choice.bind("<<ComboboxSelected>>", self.toggle_sqlite_choice)
        self.output_choice.grid(row=4, column=1, sticky="w")
        self.output_choice.current(0)
        self.sqllabel = tk.Label(self.master, text="Enter table name:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.tablename = tk.Entry(self.master, width=10)
        
        
        # Cleaning options
        tk.Label(self.master, text="Cleaning options:", font=("Times","15"), fg="#5ea832", bg="#0b2838").grid(row=5, column=0, padx=5,sticky="w")
        self.clean_option1 = tk.IntVar()
        self.clean_option2 = tk.IntVar()
        self.checkbox1 = tk.Checkbutton(self.master, text="Drop Duplicates", onvalue=1, offvalue=0, variable=self.clean_option1, bg="#0b2838")
        self.checkbox2 = tk.Checkbutton(self.master, text="Drop empty rows", onvalue=1, offvalue=0, variable=self.clean_option2, command=self.toggle_how, bg="#0b2838")
        self.checkbox1.grid(row=5, column=1, sticky="w")
        self.checkbox2.grid(row=6, column=1, sticky="w")
        
        # DropNA How section
        self.howlabel = tk.Label(self.master, text="How to filter?", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.how = tk.StringVar()
        self.checkbox2_how = ttk.Combobox(self.master, state="readonly", values=["any", "all"], textvariable=self.how, width=5)
        self.checkbox2_how.current(0)

        # Clean+Save button
        self.clean_button = tk.Button(self.master, text="Clean!", command=self.clean, activeforeground="blue")
        self.clean_button.grid(row=7, column=0, columnspan=3, padx=10, pady=10)

    ### Functions for the cleaner page ###

    ## Toggle yes for checkbox fucntion
    def toggle_index(self):
        if self.index_check.get():
            self.master.geometry("625x240")
            self.index.config(text="(yes)")
            self.index_col_label.grid(row=3, column=2, sticky="w")
            self.index_col.grid(row=3, column=3, sticky="w")
        elif not self.index_check.get() and (self.output_choice.get() == ".sqlite" or self.clean_option2.get()):
            self.master.geometry("625x240")
            self.index.config(text="(no)")
            self.index_col_label.grid_forget()
            self.index_col.grid_forget()
        else:
            self.master.geometry("520x240")
            self.index.config(text="(no)")
            self.index_col_label.grid_forget()
            self.index_col.grid_forget()

    ## Toggle SQLite choice function
    def toggle_sqlite_choice(self, event):
        if self.output_choice.get() == ".sqlite":
            self.master.geometry("625x240")
            self.sqllabel.grid(row=4, column=2, sticky="w")
            self.tablename.grid(row=4, column=3,sticky="w")
        elif self.output_choice.get() != ".sqlite" and (self.clean_option2.get() or self.index_check.get()):
            self.master.geometry("625x240")
            self.sqllabel.grid_forget()
            self.tablename.grid_forget()
        else:
            self.master.geometry("520x240")
            self.sqllabel.grid_forget()
            self.tablename.grid_forget()


    ## Toggle how section function
    def toggle_how(self):
        if self.clean_option2.get():
            self.master.geometry("625x240")
            self.howlabel.grid(row=6, column=2, sticky="w")
            self.checkbox2_how.grid(row=6, column=3, sticky="w")
        elif not self.clean_option2.get() and (self.output_choice.get() == ".sqlite" or self.index_check.get()):
            self.master.geometry("625x240")
            self.howlabel.grid_forget()
            self.checkbox2_how.grid_forget()
        else:
            self.master.geometry("520x240")
            self.howlabel.grid_forget()
            self.checkbox2_how.grid_forget()

    ## Choosing an Excel file function
    def choose_file(self):
        file_choice = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_choice:
            self.file.config(state="normal")
            self.file.delete(0, tk.END)
            self.file.insert(0, file_choice)
            self.file.config(state="disabled")

    ## SQLite handling function
    def sqlite_convert(self, filename, tablename, df):
        ###creating sqlite database
        if tablename == "":
            table_name = "my_table"
        else:
            table_name = tablename
        try:
            connection = sqlite3.connect(filename)
        except Exception as e:
            messagebox.showerror('Error', f"Error occured while creating the SQLite database: {e}")
            return
        ###converting pandas df to sqlite table
        try:
            if self.index_check.get():
                df.to_sql(table_name, connection, if_exists='replace', index=True)
            else:
                df.to_sql(table_name, connection, if_exists='replace', index=False)
            messagebox.showinfo("Success", f"Excel file converted to an SQLite database with a table \"{table_name}\" successfully!")
        except Exception as e:
            messagebox.showerror('Error', f"Error occured while converting the excel file to SQLite table: {e}")
            return
        finally:
            connection.commit()
            connection.close()

    ## Cleaning+Saving function
    def clean(self):
        if not self.file.get():
            self.file_label.config(text="* Choose Excel file:")
            self.file_label.config(fg="red")
            messagebox.showwarning("Warning", "Please choose a file!")
            return
        else:
            self.file_label.config(text="Choose Excel file:")
            self.file_label.config(fg="#5ea832")

            # Throwing an error if the directory doesn't exitst
            if not os.path.exists(self.file.get()):
                messagebox.showerror("Error", "The directory or file does not exist!")
                self.file_label.config(text="* Choose Excel file:")
                self.file_label.config(fg="red")
                return

            # Reading the original file and creating a new file with _cleaned appended and the chosen output type
            original_file_path = self.file.get()
            directory, filename = os.path.split(original_file_path)
            file_name_without_ext, file_ext = os.path.splitext(filename)
            cleaned_filename = f"{file_name_without_ext}_cleaned{self.output_choice.get()}"
            cleaned_file_path = os.path.join(directory, cleaned_filename)

            # Setting up pandas dataframe
            if self.index_check.get() == 0:
                df = pd.read_excel(original_file_path)
            else:
                df = pd.read_excel(original_file_path, index_col=int(self.index_col.get())-1)
            
            # Cleaning
            removed_count = 0
            if self.clean_option1.get():
                removed_count += df.duplicated().sum()
                df.drop_duplicates(inplace=True)
                df.reset_index(drop=True, inplace=True)
                df.index = df.index + 1
            if self.clean_option2.get():
                if self.how.get() == "any":
                    removed_count += df.isna().any().sum()
                else:
                    removed_count += df.isna().all().sum()
                df.dropna(how=self.how.get(), inplace=True)
                df.reset_index(drop=True, inplace=True)
                df.index = df.index + 1
            messagebox.showinfo("Success", f"{removed_count} rows removed.")

            # Saving file
            if self.output_choice.get() == ".xlsx":
                if self.index_check.get():
                    df.to_excel(cleaned_file_path, index=True)
                else:
                    df.to_excel(cleaned_file_path, index=False)
                messagebox.showinfo("Success", f"Excel file cleaned and saved as \"{cleaned_filename}\" successfully!")
            elif self.output_choice.get() == ".csv":
                if self.index_check.get():
                    df.to_csv(cleaned_file_path, index=True)
                else:
                    df.to_csv(cleaned_file_path, index=False)
                messagebox.showinfo("Success", f"Excel file converted to a CSV file and saved as \"{cleaned_filename}\" successfully!")
            elif self.output_choice.get() == ".sqlite":
                self.sqlite_convert(cleaned_file_path, self.tablename.get(), df)



    ## Pivot page
    def pivot_page(self):
        
        # Destroying all widgets from the main page
        for i in self.master.winfo_children():
            i.destroy()
        
        # Main geometry
        self.master.geometry("520x240")
        
        # Back button
        back_button = tk.Button(self.master, text="<<<", command=self.main_page, cursor="hand2", activeforeground="blue")
        back_button.grid(row=0, column=0, padx=10, sticky="w")

        # Pivot page title
        tk.Label(self.master, text = "Excel Pivot table Creator", font=("Times","20"), fg="#5ea832", bg="#0b2838").grid(row=1, column=1, padx=10, sticky="w")

        # Choosing an Excel file
        self.file_label = tk.Label(self.master, text="Choose Excel file:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.file_label.grid(row=2, column=0, padx=5, sticky="w")
        self.file = tk.Entry(self.master, state="disabled")
        self.file.grid(row=2, column=1, sticky="w")
        self.file_button = tk.Button(self.master, text="Choose file", command=self.choose_file, activeforeground="blue", bg="#0b2838")
        self.file_button.grid(row=2, column=2, sticky="w")

        # Choosing a data sheet
        self.sheet_label = tk.Label(self.master, text="Data Sheet name:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.sheet_label.grid(row=3, column=0, padx=5, sticky="w")
        self.sheet_choice = tk.Entry(self.master, width=10)
        self.sheet_choice.insert(0, "Sheet1")
        self.sheet_choice.grid(row=3, column=1, sticky="w")

        # choosing pivot index
        self.index_label = tk.Label(self.master, text="Pivot index column:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.index_label.grid(row=4, column=0, padx=5, sticky="w")
        self.index_choice = tk.Entry(self.master, width=10)
        self.index_choice.grid(row=4, column=1, sticky="w")

        # choosing pivot value
        self.value_label = tk.Label(self.master, text="Pivot values column:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.value_label.grid(row=5, column=0, padx=5, sticky="w")
        self.value_choice = tk.Entry(self.master, width=10)
        self.value_choice.grid(row=5, column=1, sticky="w")

        # choosing value aggregation
        tk.Label(self.master, text="Value aggregation:", font=("Times","15"), fg="#5ea832", bg="#0b2838").grid(row=6, column=0, padx=5, sticky="w")
        self.agg_choice = tk.StringVar()
        self.agg_choice_combobox = ttk.Combobox(self.master, textvariable=self.agg_choice, state="readonly", values=["sum", "mean", "max", "min", "count"], width=5)
        self.agg_choice_combobox.grid(row=6, column=1, sticky="w")
        self.agg_choice_combobox.current(0)

        # create pivot table button
        self.pivot_button = tk.Button(self.master, text="Create Pivot Table", command=self.create_pivot, activeforeground="blue", bg="#0b2838")
        self.pivot_button.grid(row=7, column=1)

    # Pivot creation and saving function
    def create_pivot(self):
        # Checking if all fields are filled
        if not self.file.get():
            self.file_label.config(text="* Choose Excel file:")
            self.file_label.config(fg="red")
            messagebox.showwarning("Warning", "Please choose a file!")
            return
        else:
            self.file_label.config(text="Choose Excel file:")
            self.file_label.config(fg="#5ea832")
            
            # Creating a new sheet for the pivot table
            wb = load_workbook(self.file.get())
            pivot_sheet = wb.create_sheet(title="PivotTable")
            
            # Reading the data from the Excel file
            try:
                data = pd.read_excel(self.file.get(), sheet_name=self.sheet_choice.get())
                self.sheet_label.config(fg="#5ea832")
                self.sheet_label.config(text="Data Sheet name:")
            except Exception:
                messagebox.showerror("Error", "Please choose a valid sheet name!")
                self.sheet_label.config(fg="red")
                self.sheet_label.config(text="* Data Sheet name:")
                return
            
            # handling missing and incorrect data
            if self.value_choice.get() not in data.columns:
                self.value_label.config(fg="red")
                self.value_label.config(text="* Pivot values column:")
                if self.index_choice.get() not in data.columns:
                    messagebox.showerror("Error", "Please choose a valid index and value column name!")
                    self.index_label.config(fg="red")
                    self.index_label.config(text="* Pivot index column:")
                    return
                self.index_label.config(fg="#5ea832")
                self.index_label.config(text="Pivot index column:")
                messagebox.showerror("Error", "Please choose a valid value column name!")
                return
            elif self.index_choice.get() not in data.columns:
                self.index_label.config(fg="red")
                self.index_label.config(text="* Pivot index column:")
                if self.value_label.get() not in data.columns:
                    messagebox.showerror("Error", "Please choose a valid index and value column name!")
                    self.value_label.config(fg="red")
                    self.value_label.config(text="* Pivot values column:")
                    return
                self.value_label.config(fg="#5ea832")
                self.value_label.config(text="Pivot value column:")
                messagebox.showerror("Error", "Please choose a valid index column name!")
                return
            else:
                if data[self.value_choice.get()].dtype == "int64" or data[self.value_choice.get()].dtype == "float64":
                    self.value_label.config(fg="#5ea832")
                    self.value_label.config(text="Pivot values column:")
                    self.index_label.config(fg="#5ea832")
                    self.index_label.config(text="Pivot index column:")
                else:
                    messagebox.showerror("Error", f"Value column has \"{data[self.value_choice.get()].dtype}\" datatype. Please choose a column with integers!")
                    self.value_label.config(fg="red")
                    self.value_label.config(text="* Pivot values column:")
                    return

            # Creating the pivot table
            pivot_table_df = pd.pivot_table(data, index=self.index_choice.get(), values=self.value_choice.get(), aggfunc=self.agg_choice.get())
            self.index_label.config(fg="#5ea832")
            self.index_label.config(text="Pivot index column:")
            self.value_label.config(fg="#5ea832")
            self.value_label.config(text="Pivot values column:")
            
            # Writing the pivot table to the Excel file
            for row in dataframe_to_rows(pivot_table_df, index=True, header=True):
                pivot_sheet.append(row)
            
            # Saving the Excel file
            wb.save(self.file.get())
            messagebox.showinfo("Success", f"Pivot table created successfully!")
            

    ## Graph plotter page
    def plotter_page(self):
        # Destroying all widgets from the main page
        for i in self.master.winfo_children():
            i.destroy()
        
        # Main geometry
        self.master.geometry("520x240")
        
        # Back button
        back_button = tk.Button(self.master, text="<<<", command=self.main_page, cursor="hand2", activeforeground="blue")
        back_button.grid(row=0, column=0, padx=10, sticky="w")

        # Graph plotter page title
        tk.Label(self.master, text = "Excel Graph Plotter", font=("Times","20"), fg="#5ea832", bg="#0b2838").grid(row=1, column=0, columnspan=3, padx=10, pady=20)

        # Choosing an Excel file and reading it into pandas dataframe
        self.file_label = tk.Label(self.master, text="Choose Excel file:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.file_label.grid(row=2, column=0, padx=5, sticky="w")
        self.file = tk.Entry(self.master, state="disabled")
        self.file.grid(row=2, column=1, sticky="w")
        self.file_button = tk.Button(self.master, text="Choose file", command=self.choose_file, activeforeground="blue", bg="#0b2838")
        self.file_button.grid(row=2, column=2)


         # Specifying if there is an index in the excell file
        tk.Label(self.master, text="Does your file have index?", font=("Times","15"), fg="#5ea832", bg="#0b2838").grid(row=3, column=0, padx=5, sticky="w")
        self.index_check = tk.IntVar()
        self.index = tk.Checkbutton(self.master, variable=self.index_check, text="(no)", onvalue=1, offvalue=0, activeforeground="blue", command=self.index_read, fg="#5ea832", bg="#0b2838")
        self.index.grid(row=3, column=1, sticky="w")
        self.index_col_label = tk.Label(self.master, text="Index num:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.index_col = tk.Entry(self.master, width=2)

        # Choosing a data sheet
        self.sheet_label = tk.Label(self.master, text="Data Sheet name:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.sheet_label.grid(row=4, column=0, padx=5, sticky="w")
        self.sheet_choice = tk.Entry(self.master, width=10)
        self.sheet_choice.insert(0, "Sheet1")
        self.sheet_choice.grid(row=4, column=1, sticky="w")

        # Choosing graph type
        tk.Label(self.master, text="Graph type:", font=("Times","15"), fg="#5ea832", bg="#0b2838").grid(row=5, column=0, padx=5, sticky="w")
        self.graph_choice = tk.StringVar()
        self.graph_choice_combobox = ttk.Combobox(self.master, textvariable=self.graph_choice, state="readonly", values=["box", "line", "bar", "scatter", "pie"], width=10)
        self.graph_choice_combobox.grid(row=5, column=1, sticky="w")
        self.graph_choice_combobox.bind("<<ComboboxSelected>>", self.graph_type_selected)
        
        # General Properties for Plotting + Plotting Button
        self.graph_label = tk.Label(self.master, text="<3", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.include_label = tk.Label(self.master, text="Include:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.include_median_var = tk.IntVar()
        self.include_median = tk.Checkbutton(self.master, variable=self.include_median_var, text="median line", onvalue=1, offvalue=0, activeforeground="blue", bg="#0b2838")
        self.include_mean_var = tk.IntVar()
        self.include_mean = tk.Checkbutton(self.master, variable=self.include_mean_var, text="mean line", onvalue=1, offvalue=0, activeforeground="blue", bg="#0b2838")
        self.graph_plot_button = tk.Button(self.master, text="Plot!", command=self.plot_graph, activeforeground="blue", bg="#0b2838", width=20)


        # Box graph parameter choosing:
        self.box_graph_param_var = tk.StringVar()
        self.box_graph_param_choice = ttk.Combobox(self.master, textvariable=self.box_graph_param_var, state="readonly", values=[None], width=10)

        # Line graph parameter choosing:
        self.line_graph_param_var = tk.StringVar()
        self.line_graph_param_choice = ttk.Combobox(self.master, textvariable=self.line_graph_param_var, state="readonly", values=[None], width=10)
        
        # Bar graph parameter choosing:
        self.bar_x_label= tk.Label(self.master, text="X axis:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.bar_x_param_var = tk.StringVar()
        self.bar_x_param_choice = ttk.Combobox(self.master, textvariable=self.bar_x_param_var, state="readonly", values=[None], width=10)
        self.bar_x_param_choice.bind("<<ComboboxSelected>>", self.bar_x_selected)
        self.bar_y_label= tk.Label(self.master, text="Y axis:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.bar_y_param_var = tk.StringVar()
        self.bar_y_param_choice = ttk.Combobox(self.master, textvariable=self.bar_y_param_var, state="readonly", values=[None], width=10)
        self.bar_y_param_choice.bind("<<ComboboxSelected>>", self.bar_y_selected)
        self.bar_agg_label= tk.Label(self.master, text="Agg", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.bar_agg_var = tk.StringVar()
        self.bar_agg_choice = ttk.Combobox(self.master, textvariable=self.bar_agg_var, state="readonly", values=["none", "sum", "count", "mean", "max", "min", "median"], width=5)
        self.bar_agg_choice.current(0)

        # Scatter graph parameter choosing:
        self.scatter_x_label= tk.Label(self.master, text="X axis:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.scatter_x_param_var = tk.StringVar()
        self.scatter_x_param_choice = ttk.Combobox(self.master, textvariable=self.scatter_x_param_var, state="readonly", values=[None], width=10)
        self.scatter_y_label= tk.Label(self.master, text="Y axis:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.scatter_y_param_var = tk.StringVar()
        self.scatter_y_param_choice = ttk.Combobox(self.master, textvariable=self.scatter_y_param_var, state="readonly", values=[None], width=10)
        self.scatter_corr_var = tk.IntVar()
        self.scatter_corr_check = tk.Checkbutton(self.master, variable=self.scatter_corr_var, text="Include correlation", onvalue=1, offvalue=0, activeforeground="blue", bg="#0b2838")

        # Pie graph parameter choosing:
        self.pie_x_label= tk.Label(self.master, text="X axis:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.pie_x_param_var = tk.StringVar()
        self.pie_x_param_choice = ttk.Combobox(self.master, textvariable=self.pie_x_param_var, state="readonly", values=[None], width=10)
        self.pie_y_label= tk.Label(self.master, text="Y axis:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.pie_y_param_var = tk.StringVar()
        self.pie_y_param_choice = ttk.Combobox(self.master, textvariable=self.pie_y_param_var, state="readonly", values=[None], width=10)
        self.pie_y_param_choice.bind("<<ComboboxSelected>>", self.pie_y_selected)
        self.pie_agg_label= tk.Label(self.master, text="Agg", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.pie_agg_var = tk.StringVar()
        self.pie_agg_choice = ttk.Combobox(self.master, textvariable=self.pie_agg_var, state="readonly", values=["sum", "count", "mean", "max", "min", "median"], width=5)
    
    # changing the input entrys and buttons according to the graph type
    def graph_type_selected(self, event):
        try:
            # Reading the excel file into pandas dataframe
            if self.index_check.get():
                self.df = pd.read_excel(self.file.get(), sheet_name=self.sheet_choice.get(), index_col=int(self.index_col.get())-1)
            else:
                self.df = pd.read_excel(self.file.get(), sheet_name=self.sheet_choice.get())
           
            # Placing the Graph Label for the graph entry
            self.graph_label.grid(row=6, column=0, padx=5, sticky="w")
            # Restoring file_label to its original state if an exception ran
            self.file_label.config(text="Choose Excel file:")
            self.file_label.config(fg="#5ea832")
            self.index_col_label.config(text="Index num:")
            self.index_col_label.config(fg="#5ea832")

            # Graph choice
            choice = self.graph_choice.get()
            # Box graph parameter choosing:
            if choice == "box":
                # label change and geometry
                self.graph_label.config(text="Box plot:")
                self.master.geometry("520x300")
                # forgetting
                self.line_graph_param_choice.grid_forget()
                self.bar_x_param_choice.grid_forget()
                self.bar_y_param_choice.grid_forget()
                self.bar_x_label.grid_forget()
                self.bar_y_label.grid_forget()
                self.bar_agg_label.grid_forget()
                self.bar_agg_choice.grid_forget()
                self.scatter_x_param_choice.grid_forget()
                self.scatter_y_param_choice.grid_forget()
                self.scatter_x_label.grid_forget()
                self.scatter_y_label.grid_forget()
                self.scatter_corr_check.grid_forget()
                self.pie_x_label.grid_forget()
                self.pie_y_label.grid_forget()
                self.pie_agg_label.grid_forget()
                self.pie_x_param_choice.grid_forget()
                self.pie_y_param_choice.grid_forget()
                self.pie_agg_choice.grid_forget()
                # setting up
                box_values = [x for x in self.df.columns if self.df[x].dtype == "int64" or self.df[x].dtype == "float64"]
                self.box_graph_param_choice.config(values=box_values)
                self.box_graph_param_choice.grid(row=6, column=1, sticky="w")
                self.box_graph_param_choice.current(0)
                self.include_label.grid(row=7, column=0, padx=5, sticky="w")
                self.include_median.grid(row=7, column=1, sticky="w")
                self.include_mean.grid(row=7, column=1, columnspan=2)
                self.graph_plot_button.grid(row=8, column=0, columnspan=3)
            # Line graph parameter choosing:
            elif choice == "line":
                # label change and geometry
                self.graph_label.config(text="Line plot:")
                self.master.geometry("520x300")
                # forgetting
                self.box_graph_param_choice.grid_forget()
                self.bar_x_param_choice.grid_forget()
                self.bar_y_param_choice.grid_forget()
                self.bar_x_label.grid_forget()
                self.bar_y_label.grid_forget()
                self.bar_agg_label.grid_forget()
                self.bar_agg_choice.grid_forget()
                self.scatter_x_param_choice.grid_forget()
                self.scatter_y_param_choice.grid_forget()
                self.scatter_x_label.grid_forget()
                self.scatter_y_label.grid_forget()
                self.scatter_corr_check.grid_forget()
                self.pie_x_label.grid_forget()
                self.pie_y_label.grid_forget()
                self.pie_agg_label.grid_forget()
                self.pie_x_param_choice.grid_forget()
                self.pie_y_param_choice.grid_forget()
                self.pie_agg_choice.grid_forget()
                # setting up
                line_values = [x for x in self.df.columns if self.df[x].dtype == "int64" or self.df[x].dtype == "float64"]
                self.line_graph_param_choice.config(values=line_values)
                self.line_graph_param_choice.current(0)
                self.line_graph_param_choice.grid(row=6, column=1, sticky="w")
                self.include_label.grid(row=7, column=0, padx=5, sticky="w")
                self.include_median.grid(row=7, column=1, sticky="w")
                self.include_mean.grid(row=7, column=1, columnspan=2)
                self.graph_plot_button.grid(row=8, column=0, columnspan=3)
            # Bar graph parameter choosing:
            elif choice == "bar":
                # label change and geometry
                self.graph_label.config(text="Bar plot:")
                self.master.geometry("520x300")
                # forgetting
                self.include_label.grid_forget()
                self.include_median.grid_forget()
                self.include_mean.grid_forget()
                self.box_graph_param_choice.grid_forget()
                self.line_graph_param_choice.grid_forget()
                self.scatter_x_param_choice.grid_forget()
                self.scatter_y_param_choice.grid_forget()
                self.scatter_x_label.grid_forget()
                self.scatter_y_label.grid_forget()
                self.scatter_corr_check.grid_forget()
                self.pie_x_label.grid_forget()
                self.pie_y_label.grid_forget()
                self.pie_agg_label.grid_forget()
                self.pie_x_param_choice.grid_forget()
                self.pie_y_param_choice.grid_forget()
                self.pie_agg_choice.grid_forget()
                # setting up
                self.bar_values_x = [x for x in self.df.columns]
                self.bar_values_y = [x for x in self.df.columns if self.df[x].dtype == "int64" or self.df[x].dtype == "float64"]
                self.bar_x_param_choice.config(values=self.bar_values_x)
                self.bar_y_param_choice.config(values=self.bar_values_y)
                self.bar_x_label.grid(row=6, column=0, columnspan=1, padx=6)
                self.bar_y_label.grid(row=7, column=0, columnspan=1, padx=6)
                self.bar_x_param_choice.grid(row=6, column=0, columnspan=2, padx=5)
                self.bar_y_param_choice.grid(row=7, column=0, columnspan=2, padx=5)
                self.bar_agg_label.grid(row=7, column=1, columnspan=2, padx=5)
                self.bar_agg_choice.grid(row=7,column=2, sticky="w")
                self.graph_plot_button.grid(row=8, column=0, columnspan=3)
            # Scatter graph parameter choosing:
            elif choice == "scatter":
                # label change and geometry
                self.graph_label.config(text="Scatter plot:")
                self.master.geometry("520x300")
                # forgetting
                self.include_label.grid_forget()
                self.include_median.grid_forget()
                self.include_mean.grid_forget()
                self.box_graph_param_choice.grid_forget()
                self.line_graph_param_choice.grid_forget()
                self.bar_x_param_choice.grid_forget()
                self.bar_y_param_choice.grid_forget()
                self.bar_x_label.grid_forget()
                self.bar_y_label.grid_forget()
                self.bar_agg_label.grid_forget()
                self.bar_agg_choice.grid_forget()
                self.pie_x_label.grid_forget()
                self.pie_y_label.grid_forget()
                self.pie_agg_label.grid_forget()
                self.pie_x_param_choice.grid_forget()
                self.pie_y_param_choice.grid_forget()
                self.pie_agg_choice.grid_forget()
                # setting up
                self.scatter_values_init = [x for x in self.df.columns if self.df[x].dtype == "int64" or self.df[x].dtype == "float64"]
                self.scatter_x_param_choice.config(values=self.scatter_values_init)
                self.scatter_y_param_choice.config(values=self.scatter_values_init)
                self.scatter_x_param_choice.current(0)
                self.scatter_y_param_choice.current(0)
                self.scatter_x_label.grid(row=6, column=0, columnspan=1, padx=5, sticky="e")
                self.scatter_y_label.grid(row=7, column=0, columnspan=1, padx=5, sticky="e")
                self.scatter_x_param_choice.grid(row=6, column=1)
                self.scatter_y_param_choice.grid(row=7, column=1)
                self.scatter_corr_check.grid(row=7, column=2, sticky="w")
                self.graph_plot_button.grid(row=8, column=0, columnspan=3)
            
            # Pie graph parameter choosing:
            elif choice == "pie":
                # label change and geometry
                self.graph_label.config(text="Pie plot:")
                self.master.geometry("520x300")
                # forgetting
                self.include_label.grid_forget()
                self.include_median.grid_forget()
                self.include_mean.grid_forget()
                self.box_graph_param_choice.grid_forget()
                self.line_graph_param_choice.grid_forget()
                self.bar_x_param_choice.grid_forget()
                self.bar_y_param_choice.grid_forget()
                self.bar_x_label.grid_forget()
                self.bar_y_label.grid_forget()
                self.bar_agg_label.grid_forget()
                self.bar_agg_choice.grid_forget()
                self.scatter_x_param_choice.grid_forget()
                self.scatter_y_param_choice.grid_forget()
                self.scatter_x_label.grid_forget()
                self.scatter_y_label.grid_forget()
                self.scatter_corr_check.grid_forget()
                # setting up
                self.pie_values_init = [x for x in self.df.columns]
                self.pie_x_param_choice.config(values=self.pie_values_init)
                self.pie_y_param_choice.config(values=self.pie_values_init)
                self.pie_x_label.grid(row=6, column=0, columnspan=1, padx=5, sticky="e")
                self.pie_y_label.grid(row=7, column=0, columnspan=1, padx=5, sticky="e")
                self.pie_x_param_choice.grid(row=6, column=1, sticky="w")
                self.pie_y_param_choice.grid(row=7, column=1, sticky="w")
                self.pie_agg_label.grid(row=7, column=2, sticky="w")
                self.pie_agg_choice.grid(row=7, column=2, sticky="e")
                self.graph_plot_button.grid(row=8, column=0, columnspan=3)
        # exception for the case when there is no file chosen
        except Exception:
            if not os.path.exists(self.file.get()):
                self.file_label.config(text="* Choose Excel file:")
                self.file_label.config(fg="red")
                tk.messagebox.showerror("Error", "Please choose an Excel file first!")
            if self.index_check.get() == 1 and self.index_col.get() == "":
                self.index_col_label.config(text="* Index num:")
                self.index_col_label.config(fg="red")
                tk.messagebox.showerror("Error", "Please specify an index first!")
            return

    # Handling the choices of parameters in bar plot.
    def bar_x_selected(self, event):
        choice = self.bar_x_param_var.get()
        val = [x for x in self.df.columns if (self.df[x].dtype == "int64" or self.df[x].dtype == "float64") and x != choice]
        self.bar_y_param_choice.config(values=val)
    def bar_y_selected(self, event):
        choice = self.bar_y_param_var.get()
        val = [x for x in self.df.columns if x != choice]
        self.bar_x_param_choice.config(values=val)
    def pie_y_selected(self, event):
        if self.df[self.pie_y_param_var.get()].dtype not in ["int64", "float64", "int32", "float32"]:
            self.pie_agg_choice.config(values=["count"])
        else:
            self.pie_agg_choice.config(values=["sum", "count", "mean", "max", "min", "median"])
    
    # Toggle the index column function
    def index_read(self):
        if self.index_check.get() == 1:
            self.index_col_label.grid(row=3, column=1,columnspan=2)
            self.index_col.grid(row=3, column=1,columnspan=2, padx=70, sticky="e")
            self.index.config(text="(yes)")
        else:
            self.index_col_label.grid_forget()
            self.index_col.grid_forget()
            self.index.config(text="(no)")

    # Plotting a graph function
    def plot_graph(self):
        # Box plot
        if self.graph_choice.get() == "box":
            try:
                plotting_data = self.df[self.box_graph_param_var.get()]
                plotting_data.plot(kind="box", vert=False)
                plt.title(self.box_graph_param_var.get())
                if self.include_median_var.get() == 1:
                    plt.axvline(plotting_data.median(), color="green", label=f"median({round(plotting_data.median(),2)})")
                    plt.legend(loc="best")
                if self.include_mean_var.get() == 1:
                    plt.axvline(plotting_data.mean(), color="red", linestyle="-.", label=f"mean({round(plotting_data.mean(),2)})")
                    plt.legend(loc="best")
            except KeyError:
                tk.messagebox.showerror("Error", "The column you chose has missing values! Plotting the graph without the NaN rows.")
                plotting_data = self.df[self.df[self.box_graph_param_var.get()] != np.nan]
                plotting_data.plot(kind="box", vert=False)
                plt.title(self.box_graph_param_var.get())
                if self.include_median_var.get() == 1:
                    plt.axvline(plotting_data.median(), color="green", label=f"median({round(plotting_data.median(),2)})")
                    plt.legend(loc="best")
                if self.include_mean_var.get() == 1:
                    plt.axvline(plotting_data.mean(), color="red", linestyle="-.", label=f"mean({round(plotting_data.mean(),2)})")
                    plt.legend(loc="best")
        # Line plot
        elif self.graph_choice.get() == "line":
            try:
                plotting_data = self.df[self.line_graph_param_var.get()]
                plotting_data.plot(kind="line")
                plt.title(self.line_graph_param_var.get())
                if self.include_median_var.get() == 1:
                    plt.axhline(plotting_data.median(), color="green", label=f"median({round(plotting_data.median(),2)})")
                    plt.legend(loc="best")
                if self.include_mean_var.get() == 1:
                    plt.axhline(plotting_data.mean(), color="red", linestyle="-.", label=f"mean({round(plotting_data.mean(),2)})")
                    plt.legend(loc="best")
            except KeyError:
                tk.messagebox.showerror("Error", "The column you chose has missing values! Plotting the graph without the NaN rows.")
                plotting_data = self.df[self.df[self.line_graph_param_var.get()] != np.nan]
                plotting_data.plot(kind="line")
                plt.title(self.line_graph_param_var.get())
                if self.include_median_var.get() == 1:
                    plt.axhline(plotting_data.median(), color="green", label=f"median({round(plotting_data.median(),2)})")
                    plt.legend(loc="best")
                if self.include_mean_var.get() == 1:
                    plt.axhline(plotting_data.mean(), color="red", linestyle="-.", label=f"mean({round(plotting_data.mean(),2)})")
                    plt.legend(loc="best")
        # Bar plot
        elif self.graph_choice.get() == "bar":
            try:
                agg = self.bar_agg_choice.get()
                self.df.groupby(self.bar_x_param_choice.get())[self.bar_y_param_choice.get()].agg(agg).plot(kind="bar")
                plt.title(f"{self.bar_x_param_choice.get()} by {self.bar_y_param_choice.get()} {agg}")
            except KeyError:
                if self.bar_x_param_choice.get() == "":
                    tk.messagebox.showerror("Error", "You have to choose a column for the x axis!")
                    self.bar_x_label.config(fg="red")
                elif self.bar_y_param_choice.get() == "":
                    tk.messagebox.showerror("Error", "You have to choose a column for the y axis!")
                    self.bar_y_label.config(fg="red")
            except AttributeError:
                tk.messagebox.showerror("Error", "Something went wrong!")
        # Scatter plot
        elif self.graph_choice.get() == "scatter":
            try:
                if self.scatter_corr_var.get() == 1:
                    self.df.plot(kind="scatter", x=self.scatter_x_param_choice.get(), y=self.scatter_y_param_choice.get())
                    legend_label=[f"Correlation: {round(self.df[self.scatter_x_param_choice.get()].corr(self.df[self.scatter_y_param_choice.get()]),2)}"]
                    plt.legend(legend_label,loc="best")
                else:
                    self.df.plot(kind="scatter", x=self.scatter_x_param_choice.get(), y=self.scatter_y_param_choice.get())
                plt.title(f"{self.scatter_x_param_choice.get()} by {self.scatter_y_param_choice.get()}")
            except KeyError:
                if self.scatter_x_param_choice.get() == "":
                    tk.messagebox.showerror("Error", "You have to choose a column for the x axis!")
                    self.scatter_x_label.config(fg="red")
                elif self.scatter_y_param_choice.get() == "":
                    tk.messagebox.showerror("Error", "You have to choose a column for the y axis!")
                    self.scatter_y_label.config(fg="red")
        # Pie plot
        elif self.graph_choice.get() == "pie":
            try:
                agg = self.pie_agg_var.get()
                self.df.groupby(self.pie_x_param_choice.get())[self.pie_y_param_choice.get()].agg(agg).plot(kind="pie")
                legend_labels = [f"{index}: {round(value,2)}" for index, value in self.df.groupby(self.pie_x_param_choice.get())[self.pie_y_param_choice.get()].agg(agg).items()]
                plt.title(f"{self.pie_x_param_choice.get()} by {self.pie_y_param_choice.get()} {agg}")
                plt.legend(legend_labels, loc="best")
                plt.ylabel("")
            except Exception:
                tk.messagebox.showerror("Error", "Something went wrong!")
                return
        try:
            # creating a new sheet for the graph
            wb = load_workbook(self.file.get())
            new_sheet = wb.create_sheet("EM Graph")

            #handling os file path
            original_file_path = self.file.get()
            directory, filename = os.path.split(original_file_path)
            # creating a graph folder
            if not os.path.exists(os.path.join(directory, "Graphs")):
                os.makedirs(os.path.join(directory, "Graphs"))

            image_filename = f"{os.path.splitext(filename)[0]}_{self.graph_choice.get()}.png"
            image_file_path = os.path.join(directory, "Graphs", image_filename)
        
            plt.savefig(image_file_path, format="png", bbox_inches="tight")
            messagebox.showinfo("Success", f"Graph saved at {image_file_path} and new sheet in the Excel file.")
            # saving the graph
            new_sheet.add_image(Image(image_file_path), anchor="A1")
            wb.save(self.file.get())
            wb.close()
            plt.close()
        
        except Exception:
            tk.messagebox.showerror("Error", "Something went wrong!")
            plt.close()
            return
        
            

        



    ## Formula applier page 
    def formula_page(self):
        # Destroying all widgets from the main page
        for i in self.master.winfo_children():
            i.destroy()
        
        # Main geometry
        self.master.geometry("520x240")
        
        # Back button
        back_button = tk.Button(self.master, text="<<<", command=self.main_page, cursor="hand2", activeforeground="blue")
        back_button.grid(row=0, column=0, padx=10, sticky="w")

        # Fromula applier page title
        tk.Label(self.master, text = "Excel Formula Applier", font=("Times","20"), fg="#5ea832", bg="#0b2838").grid(row=1, column=1, columnspan=2, pady=20)

        # Choosing an Excel file and reading it into pandas dataframe
        self.file_label = tk.Label(self.master, text="Choose Excel file:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.file_label.grid(row=2, column=0, padx=5, sticky="w")
        self.file = tk.Entry(self.master, state="disabled")
        self.file.grid(row=2, column=1, sticky="w")
        self.file_button = tk.Button(self.master, text="Choose file", command=self.choose_file, activeforeground="blue", bg="#0b2838")
        self.file_button.grid(row=2, column=2)

        # Choosing a data sheet
        self.sheet_label = tk.Label(self.master, text="Data Sheet name:", font=("Times","15"), fg="#5ea832", bg="#0b2838")
        self.sheet_label.grid(row=3, column=0, padx=5, sticky="w")
        self.sheet_choice = tk.Entry(self.master, width=10)
        self.sheet_choice.insert(0, "Sheet1")
        self.sheet_choice.grid(row=3, column=1, sticky="w")

#### GUI ####
gui = tk.Tk()
#### App ####
app = App(gui)
#### App Main Loop ####
app.master.mainloop()
